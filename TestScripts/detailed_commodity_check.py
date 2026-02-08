#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
详细检查商品数据结构
"""

import sys
import os
# 添加项目根目录到Python路径
project_root = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.append(project_root)

from excel_manager import ExcelManager
import pandas as pd

def detailed_analysis():
    """详细分析数据结构"""
    print("详细分析商品数据结构")
    print("="*60)
    
    excel_manager = ExcelManager()
    
    # 直接用openpyxl读取，看看Excel的真实结构
    from openpyxl import load_workbook
    wb = load_workbook(excel_manager.filename)
    ws = wb["商品信息"]
    
    print("Excel文件'商品信息'表的前几行内容:")
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx < 8:  # 显示前8行
            print(f"Excel行 {row_idx+1}: {row}")
        else:
            break
    
    print(f"\n总共 {ws.max_row} 行数据")
    
    # 使用pandas读取
    df = excel_manager.get_all_commodities()
    print(f"\nPandas读取的DataFrame结构:")
    print(f"DataFrame行数: {len(df)}")
    print(f"DataFrame列数: {len(df.columns)}")
    print(f"DataFrame列名: {list(df.columns)}")
    
    print(f"\nDataFrame前几行数据:")
    for idx in range(min(5, len(df))):
        print(f"DataFrame行 {idx}: {dict(df.iloc[idx])}")
    
    print(f"\n检查T001商品:")
    if '商品编号' in df.columns:
        t001_rows = df[df['商品编号'] == 'T001']
        print(f"T001在DataFrame中的数量: {len(t001_rows)}")
        for idx, row in t001_rows.iterrows():
            print(f"  T001在DataFrame行 {idx}: {row['商品名称']}")
    
    print(f"\n使用 df.iloc[1:] 后的情况:")
    df_after_slice = df.iloc[1:]
    print(f"切片后行数: {len(df_after_slice)}")
    
    if '商品编号' in df_after_slice.columns:
        t001_after_slice = df_after_slice[df_after_slice['商品编号'] == 'T001']
        print(f"切片后T001的数量: {len(t001_after_slice)}")
        if len(t001_after_slice) == 0:
            print("  T001已丢失！因为它在被切掉的第0行")
    
    print(f"\n正确的处理方式应该是保留所有数据行，不切掉任何行")
    print(f"因为pandas.read_excel已经将Excel第一行作为列标题处理了")

def main():
    detailed_analysis()

if __name__ == "__main__":
    main()