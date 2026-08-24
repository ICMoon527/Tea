import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import shutil
from datetime import datetime
import random
from utils import convert_to_jin
from logger import get_logger


class ExcelManager:
    def __init__(self, filename="tea_inventory.xlsx"):
        self.logger = get_logger()
        self.filename = filename
        self._cache = {}
        self._dirty_flags = {}
        self._pending_writes = {}
        self._dirty = False
        self.init_excel_file()
        self.flush()
        self.clear_cache()

    def init_excel_file(self):
        """初始化Excel文件，创建必要的工作表并检查结构"""
        if not os.path.exists(self.filename):
            wb = Workbook()

            # 创建商品表
            ws_commodity = wb.active
            ws_commodity.title = "商品信息"
            ws_commodity.append([
                "商品编号", "茶类", "品种", "公司", "产区", "商品名称",
                "规格", "成本价", "零售价", "生产日期",
                "保质期(月)", "当前库存", "品质特征", "年份", "等级", "单位"
            ])

            # 创建销售记录表
            ws_sell = wb.create_sheet("销售记录")
            ws_sell.append([
                "销售编号", "商品编号", "商品名称", "销售数量",
                "单价", "应收金额", "实收金额", "客户名称", "销售日期", "销售单位", "是否作废"
            ])

            # 创建进货记录表
            ws_stock = wb.create_sheet("进货记录")
            ws_stock.append([
                "进货编号", "商品编号", "商品名称", "进货数量",
                "进货单价", "供应商", "进货日期", "备注", "进货单位"
            ])

            # 创建供应商表
            ws_supplier = wb.create_sheet("供应商")
            ws_supplier.append([
                "供应商编号", "供应商名称", "联系人", "联系电话", "地址", "累计交易金额", "备注"
            ])

            # 创建客户表
            ws_customer = wb.create_sheet("客户信息")
            ws_customer.append([
                "客户编号", "客户名称", "联系电话", "电子邮箱", "地址",
                "累计消费", "订单数", "最后购买日期", "客户等级", "备注", "创建日期"
            ])

            wb.save(self.filename)
        else:
            # 如果文件已存在，检查并修复供应商表结构
            try:
                wb = load_workbook(self.filename)
                if "供应商" in wb.sheetnames:
                    ws = wb["供应商"]
                    # 检查列数是否正确
                    if ws.max_row > 0:
                        first_row = [cell.value for cell in ws[1]]
                        expected_columns = ["供应商编号", "供应商名称", "联系人", "联系电话", "地址", "累计交易金额", "备注"]
                        if first_row != expected_columns:
                            # 需要修复供应商表结构
                            df = pd.read_excel(self.filename, sheet_name="供应商", engine='openpyxl')
                            # 添加缺失的列
                            if "累计交易金额" not in df.columns:
                                df["累计交易金额"] = 0.0
                            # 确保列顺序正确
                            df = df[expected_columns]
                            # 重写供应商表
                            self.write_sheet("供应商", df)
                wb.close()
            except (PermissionError, OSError) as e:
                self.logger.error(f"检查Excel文件结构时出错: {e}")
    
    def read_sheet(self, sheet_name: str):
        """读取指定工作表的数据（带缓存）"""
        if sheet_name in self._cache:
            return self._cache[sheet_name].copy()
        try:
            df = pd.read_excel(self.filename, sheet_name=sheet_name, engine='openpyxl')
            # 过滤掉 Unnamed 列
            df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

            # 自愈：过滤掉误写入数据区的重复表头行（避免列表显示重复表头并防止其传播）
            if not df.empty and len(df.columns) > 1:
                _first = df.columns[0]
                _rest = df.iloc[:, 1:]
                _rest_matches = _rest.eq(pd.Series(list(_rest.columns), index=_rest.columns), axis=1).sum(axis=1)
                # 表头行特征：首列等于列名，且至少一个其他单元格等于对应列名
                _header_like = (df[_first] == _first) & (_rest_matches >= 1)
                if _header_like.any():
                    df = df[~_header_like].reset_index(drop=True)
            
            # 针对特定工作表进行数据类型处理
            if sheet_name == "客户信息" and not df.empty:
                # 确保累计消费列是浮点类型
                if '累计消费' in df.columns:
                    df['累计消费'] = pd.to_numeric(df['累计消费'], errors='coerce').fillna(0.0)
            elif sheet_name == "供应商" and not df.empty:
                # 确保累计交易金额列是浮点类型
                if '累计交易金额' in df.columns:
                    df['累计交易金额'] = pd.to_numeric(df['累计交易金额'], errors='coerce').fillna(0.0)
            elif sheet_name == "商品信息" and not df.empty:
                # 确保数值列是浮点类型
                numeric_columns = ['成本价', '零售价', '当前库存']
                for col in numeric_columns:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
            elif sheet_name == "销售记录" and not df.empty:
                # 将是否作废统一转换为布尔值，兼容历史数据中的字符串形式（TRUE/FALSE/true/false）
                if '是否作废' in df.columns:
                    df['是否作废'] = df['是否作废'].map(
                        lambda v: str(v).strip().lower() == 'true' if pd.notna(v) else False
                    )
            
            self._cache[sheet_name] = df.copy()
            self._dirty_flags[sheet_name] = False
            return df
        except (FileNotFoundError, PermissionError, ValueError, OSError) as e:
            self.logger.error(f"读取工作表 {sheet_name} 出错: {e}")
            return pd.DataFrame()
    
    def clear_cache(self, sheet_name: str = None):
        """清空所有缓存数据"""
        if sheet_name is None:
            self.flush()
        self._cache.clear()
        self._dirty_flags.clear()
        if sheet_name is None:
            self._pending_writes.clear()
            self._dirty = False
    
    def write_sheet(self, sheet_name: str, data):
        """写入数据到指定工作表（延迟写入）"""
        expected_columns = {
            "商品信息": ["商品编号", "茶类", "品种", "公司", "产区", "商品名称", "规格", "成本价", "零售价", "生产日期", "保质期(月)", "当前库存", "品质特征", "年份", "等级", "单位"],
            "销售记录": ["销售编号", "商品编号", "商品名称", "销售数量", "单价", "应收金额", "实收金额", "客户名称", "销售日期", "销售单位", "是否作废"],
            "进货记录": ["进货编号", "商品编号", "商品名称", "进货数量", "进货单价", "供应商", "进货日期", "备注", "进货单位"],
            "供应商": ["供应商编号", "供应商名称", "联系人", "联系电话", "地址", "累计交易金额", "备注"],
            "客户信息": ["客户编号", "客户名称", "联系电话", "电子邮箱", "地址", "累计消费", "订单数", "最后购买日期", "客户等级", "备注", "创建日期"]
        }

        data_to_write = data.copy() if not data.empty else data

        if sheet_name == "商品信息" and not data_to_write.empty and '当前库存' in data_to_write.columns:
            data_to_write['当前库存'] = pd.to_numeric(data_to_write['当前库存'], errors='coerce').fillna(0).round(2)

        if not data_to_write.empty:
            columns_to_write = data_to_write.columns.tolist()
        elif sheet_name in expected_columns:
            columns_to_write = expected_columns[sheet_name]
            data_to_write = pd.DataFrame(columns=columns_to_write)
        else:
            columns_to_write = []
            data_to_write = pd.DataFrame(columns=columns_to_write)

        self._pending_writes[sheet_name] = data_to_write
        self._dirty = True
        self._cache[sheet_name] = data_to_write.copy()
        self._dirty_flags[sheet_name] = False
    
    def append_to_sheet(self, sheet_name, data_row):
        """向指定工作表追加一行数据（更新缓存）"""
        self.flush()
        try:
            wb = load_workbook(self.filename)
            ws = wb[sheet_name]

            if isinstance(data_row, pd.DataFrame):
                for _, row in data_row.iterrows():
                    ws.append(row.tolist())
            else:
                ws.append(data_row)

            wb.save(self.filename)
            if sheet_name in self._cache:
                del self._cache[sheet_name]
                del self._dirty_flags[sheet_name]
        except (PermissionError, OSError) as e:
            self.logger.error(f"追加数据到工作表 {sheet_name} 出错: {e}")
    
    def generate_id(self, prefix: str, sheet_name: str = None, id_column: str = None):
        """生成唯一ID"""
        # 如果是商品编号（前缀为"C"），则使用递增编号
        if prefix == "C" and sheet_name == "商品信息" and id_column == "商品编号":
            df = self.read_sheet(sheet_name)
            if df.empty or id_column not in df.columns:
                # 如果工作表为空或列不存在，返回T001
                return "T001"
            
            # 提取现有商品编号中的数字部分
            existing_ids = df[id_column].astype(str).tolist()
            existing_numbers = []
            for id_val in existing_ids:
                if isinstance(id_val, str) and id_val.startswith("T"):
                    num_part = id_val[1:]  # 去掉"T"前缀
                    if num_part.isdigit():
                        existing_numbers.append(int(num_part))
            
            # 找到最大编号，返回下一个
            if existing_numbers:
                next_num = max(existing_numbers) + 1
            else:
                next_num = 1
            
            return f"T{next_num:03d}"
        
        # 如果提供了工作表名和ID列，则确保生成的ID不重复
        if sheet_name and id_column:
            max_attempts = 100  # 最大尝试次数，避免无限循环
            attempt = 0
            while attempt < max_attempts:
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                random_part = f"{random.randint(1000, 9999)}"
                new_id = f"{prefix}{timestamp}{random_part}"
                
                # 检查ID是否已存在
                df = self.read_sheet(sheet_name)
                if df.empty or id_column not in df.columns:
                    return new_id  # 如果工作表为空或列不存在，直接返回
                
                existing_ids = df[id_column].astype(str).tolist()
                if new_id not in existing_ids:
                    return new_id  # 找到不重复的ID
                
                attempt += 1
            
            raise Exception(f"无法生成唯一ID: {prefix}")  # 如果尝试失败，抛出异常
        else:
            # 如果没有提供工作表名和ID列，则使用时间戳和随机数
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            random_part = f"{random.randint(1000, 9999)}"
            return f"{prefix}{timestamp}{random_part}"
    
    # 商品相关操作
    def get_all_commodities(self):
        df = self.read_sheet("商品信息")
        if not df.empty and '当前库存' in df.columns:
            df['当前库存'] = pd.to_numeric(df['当前库存'], errors='coerce').fillna(0).round(2)
        return df
    
    def get_commodity_by_id(self, com_id):
        df = self.read_sheet("商品信息")
        if df.empty:
            return None
        result = df[df['商品编号'] == com_id]
        if not result.empty:
            # 确保当前库存是两位小数
            commodity = result.iloc[0].copy()
            if '当前库存' in commodity:
                commodity['当前库存'] = round(float(commodity['当前库存']), 2)
            return commodity
        return None
    
    def add_record(self, sheet_name: str, data):
        """通用添加记录方法"""
        self.append_to_sheet(sheet_name, data)
        return {'success': True}

    def update_record(self, sheet_name: str, record_id, updates, id_column: str):
        """通用更新记录方法"""
        df = self.read_sheet(sheet_name)
        if df.empty:
            return {'success': False, 'message': '表为空'}
        df[id_column] = df[id_column].astype(str)
        idx = df[df[id_column] == str(record_id)].index
        if len(idx) == 0:
            return {'success': False, 'message': '记录不存在'}
        for col, value in updates.items():
            df.at[idx[0], col] = value
        self.write_sheet(sheet_name, df)
        return {'success': True}

    def delete_record(self, sheet_name: str, record_id, id_column: str):
        """通用删除记录方法"""
        df = self.read_sheet(sheet_name)
        if df.empty:
            return {'success': False, 'message': '表为空'}
        df[id_column] = df[id_column].astype(str)
        df = df[df[id_column] != str(record_id)]
        self.write_sheet(sheet_name, df)
        return {'success': True}

    def add_commodity(self, commodity_data):
        self.append_to_sheet("商品信息", commodity_data)
        return {'success': True}
    
    def update_commodity(self, com_id, new_data):
        df = self.read_sheet("商品信息")
        if df.empty:
            return {'success': False, 'message': '商品表为空'}

        idx = df[df['商品编号'] == com_id].index
        if len(idx) > 0:
            for col, value in new_data.items():
                if col in ['当前库存', '成本价', '零售价']:
                    col_list = df[col].tolist()
                    if col == '当前库存':
                        col_list[idx[0]] = round(float(value), 2)
                    else:
                        col_list[idx[0]] = float(value)
                    df[col] = pd.Series(col_list, dtype=float)
                else:
                    df.at[idx[0], col] = value
            self.write_sheet("商品信息", df)
            return {'success': True}
        return {'success': False, 'message': '商品不存在'}
    
    def delete_commodity(self, com_id):
        df = self.read_sheet("商品信息")
        if df.empty:
            return {'success': False, 'message': '商品表为空'}

        df = df[df['商品编号'] != com_id]
        self.write_sheet("商品信息", df)
        return {'success': True}
    
    # 销售相关操作
    def get_all_sales(self, include_voided=False):
        """获取所有销售记录
        
        Args:
            include_voided: 是否包含作废记录，默认为 False
            
        Returns:
            销售记录 DataFrame
        """
        df = self.read_sheet("销售记录")
        
        if not include_voided and not df.empty:
            # 过滤掉作废记录
            if '是否作废' in df.columns:
                df = df[df['是否作废'] != True]
        
        # 新的销售记录排在前面（按销售日期倒序）
        if not df.empty and '销售日期' in df.columns:
            df = df.sort_values('销售日期', ascending=False, na_position='last')
        
        return df
    
    def add_sale(self, sale_data):
        self.append_to_sheet("销售记录", sale_data)
        
        # 更新库存
        commodity = self.get_commodity_by_id(sale_data[1])  # 商品编号在第二列
        if commodity is not None:
            current_stock = float(commodity['当前库存'])
            sold_qty = float(sale_data[3])  # 销售数量在第四列
            sale_unit = sale_data[9] if len(sale_data) > 9 else '克'  # 销售单位在第10列
            
            # 使用convert_to_jin函数转换
            sold_qty_jin = convert_to_jin(sold_qty, sale_unit)
            
            new_stock = current_stock - sold_qty_jin
            self.update_commodity(sale_data[1], {'当前库存': new_stock})
        
        # 更新客户信息
        customer_name = sale_data[7]  # 客户名称在第8列
        received_amount = sale_data[6]  # 实收金额在第7列
        sale_date = sale_data[8]  # 销售日期在第9列
        
        if customer_name:
            self.update_customer_after_sale(customer_name, received_amount, sale_date)
    
    # 进货相关操作
    def get_all_stocks(self):
        return self.read_sheet("进货记录")
    
    def add_stock(self, stock_data):
        self.append_to_sheet("进货记录", stock_data)
        
        # 更新库存和计算移动加权平均成本价
        commodity = self.get_commodity_by_id(stock_data[1])  # 商品编号在第二列
        if commodity is not None:
            current_stock = float(commodity['当前库存'])
            current_cost_price = float(commodity['成本价']) if pd.notna(commodity['成本价']) else 0.0
            
            stock_qty = float(stock_data[3])  # 进货数量在第四列
            stock_unit_price = float(stock_data[4])  # 进货单价在第五列
            stock_unit = stock_data[8] if len(stock_data) > 8 else '斤'  # 进货单位在第9列
            
            # 使用convert_to_jin函数转换进货数量为斤
            stock_qty_jin = convert_to_jin(stock_qty, stock_unit)
            
            # 计算进货单价（转换为每斤的价格）
            # 如果进货单位是克，进货单价需要乘以500得到每斤价格
            if stock_unit == '克':
                stock_unit_price_per_jin = stock_unit_price * 500
            else:
                stock_unit_price_per_jin = stock_unit_price
            
            # 计算移动加权平均成本价
            # 公式：(当前成本价×当前库存 + 新进货单价×新进货数量) / (当前库存 + 新进货数量)
            total_current_cost = current_cost_price * current_stock
            total_new_cost = stock_unit_price_per_jin * stock_qty_jin
            total_stock = current_stock + stock_qty_jin
            
            if total_stock > 0:
                new_cost_price = (total_current_cost + total_new_cost) / total_stock
                # 保留两位小数
                new_cost_price = round(new_cost_price, 2)
            else:
                new_cost_price = stock_unit_price_per_jin
            
            # 同时更新库存和成本价
            updates = {
                '当前库存': total_stock,
                '成本价': new_cost_price
            }
            self.update_commodity(stock_data[1], updates)
        
        # 更新供应商信息
        supplier_name = stock_data[5]  # 供应商名称在第6列
        amount = float(stock_data[3]) * float(stock_data[4])  # 数量 * 单价
        stock_date = stock_data[6]  # 进货日期在第7列
        
        if supplier_name:
            self.update_supplier_after_stock(supplier_name, amount, stock_date)
    
    # 供应商相关操作
    def get_all_suppliers(self):
        try:
            self.logger.debug("=== 开始读取供应商数据 ===")
            df = self.read_sheet("供应商")
            self.logger.debug(f"原始数据行数: {len(df)}")
            self.logger.debug(f"原始数据列名: {list(df.columns)}")
            self.logger.debug(f"原始数据是否为空: {df.empty}")
            
            # 检查数据
            if not df.empty:
                # 打印原始数据的前几行
                self.logger.debug("\n原始数据前3行:")
                for i in range(min(3, len(df))):
                    self.logger.debug(f"  行 {i+1}: {list(df.iloc[i])}")
            
            # 数据清理
            if not df.empty:
                # 1. 首先移除所有字段都是 NaN 的行
                df = df.dropna(how='all')
                self.logger.debug(f"\n移除全 NaN 行后，行数: {len(df)}")
                
                if not df.empty:
                    # 2. 只对文本列进行字符串转换和去空白
                    text_columns = ['供应商编号', '供应商名称', '联系人', '联系电话', '地址']
                    for col in text_columns:
                        if col in df.columns:
                            df[col] = df[col].astype(str).str.strip()
                    
                    # 3. 移除空行（检查关键文本字段）
                    if '供应商名称' in df.columns:
                        mask = df['供应商名称'].notna() & (df['供应商名称'] != '') & (df['供应商名称'] != 'nan') & (df['供应商名称'] != 'NaN')
                        df = df[mask]
                    
                    # 4. 重置索引
                    df = df.reset_index(drop=True)
                    self.logger.debug(f"清理后的数据行数: {len(df)}")
                    
                    if not df.empty:
                        self.logger.debug("\n清理后的数据:")
                        for index, row in df.iterrows():
                            self.logger.debug(f"  行 {index+1}: {list(row)}")
            
            self.logger.debug("=== 读取供应商数据完成 ===")
            return df
        except (FileNotFoundError, PermissionError, ValueError, KeyError, OSError) as e:
            self.logger.error(f"获取供应商数据时出错: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def add_supplier(self, supplier_data):
        self.append_to_sheet("供应商", supplier_data)

    def update_supplier_after_stock(self, supplier_name, amount, stock_date):
        """进货后更新供应商信息"""
        df = self.read_sheet("供应商")
        
        # 检查并添加累计交易金额列（如果缺失）
        if "累计交易金额" not in df.columns:
            df["累计交易金额"] = 0.0
        
        if df.empty:
            # 如果供应商表为空，创建新供应商
            self._create_new_supplier(supplier_name, amount, stock_date)
            return

        # 查找供应商
        supplier_row = df[df['供应商名称'] == supplier_name]
        if supplier_row.empty:
            # 创建新供应商
            self._create_new_supplier(supplier_name, amount, stock_date)
        else:
            # 更新现有供应商
            idx = supplier_row.index[0]
            # 获取当前累计交易金额，如果为空则为0
            current_amount = float(df.at[idx, '累计交易金额']) if pd.notna(df.at[idx, '累计交易金额']) else 0.0
            # 更新累计交易金额
            new_amount = current_amount + amount
            
            # 完全重新创建列来确保数据类型正确
            amount_list = df['累计交易金额'].tolist()
            amount_list[idx] = float(new_amount)
            df['累计交易金额'] = pd.Series(amount_list, dtype=float)
            
            # 保存更新
            self.write_sheet("供应商", df)
            self.logger.info(f"供应商 {supplier_name} 已更新，累计交易金额: {new_amount}")

    def _create_new_supplier(self, supplier_name, amount, stock_date):
        """创建新供应商"""
        # 生成供应商编号
        supplier_id = self.generate_id("SP", "供应商", "供应商编号")
        
        # 创建供应商数据（包含初始累计交易金额）
        supplier_data = [
            supplier_id, supplier_name, "", "", "", amount, ""
        ]
        
        self.append_to_sheet("供应商", supplier_data)
        self.logger.info(f"新供应商 {supplier_name} 创建成功，供应商编号: {supplier_id}，初始累计交易金额: {amount}")

    # 客户相关操作
    def get_all_customers(self):
        try:
            self.logger.debug("=== 开始读取客户数据 ===")
            df = self.read_sheet("客户信息")
            self.logger.debug(f"原始数据行数: {len(df)}")
            self.logger.debug(f"原始数据列名: {list(df.columns)}")
            self.logger.debug(f"原始数据是否为空: {df.empty}")
            
            # 检查数据
            if not df.empty:
                # 打印原始数据的前几行
                self.logger.debug("\n原始数据前3行:")
                for i in range(min(3, len(df))):
                    self.logger.debug(f"  行 {i+1}: {list(df.iloc[i])}")
            
            # 数据清理
            if not df.empty:
                # 1. 首先移除所有字段都是 NaN 的行
                df = df.dropna(how='all')
                self.logger.debug(f"\n移除全 NaN 行后，行数: {len(df)}")
                
                if not df.empty:
                    # 2. 确保累计消费列是浮点类型
                    if '累计消费' in df.columns:
                        df['累计消费'] = pd.to_numeric(df['累计消费'], errors='coerce').fillna(0.0)
                    
                    # 3. 只对文本列进行字符串转换和去空白
                    text_columns = ['客户编号', '客户名称', '联系人', '联系电话', '地址', '客户等级', '最后购买日期']
                    for col in text_columns:
                        if col in df.columns:
                            df[col] = df[col].astype(str).str.strip()
                    
                    # 4. 移除空行（检查关键文本字段）
                    if '客户名称' in df.columns:
                        mask = df['客户名称'].notna() & (df['客户名称'] != '') & (df['客户名称'] != 'nan') & (df['客户名称'] != 'NaN')
                        df = df[mask]
                    
                    # 5. 自动更新所有客户等级（根据最新的门槛）
                    if '累计消费' in df.columns and '客户等级' in df.columns:
                        need_update = False
                        for idx, row in df.iterrows():
                            total_purchases = float(row['累计消费']) if pd.notna(row['累计消费']) else 0.0
                            expected_level = self.calculate_customer_level(total_purchases)
                            if str(row['客户等级']) != expected_level:
                                df.at[idx, '客户等级'] = expected_level
                                need_update = True
                        # 如果有等级变化，保存到 Excel
                        if need_update:
                            self.write_sheet("客户信息", df)
                            self.logger.info(f"已自动更新 {len(df)} 位客户的等级")
                    
                    # 5. 重置索引
                    df = df.reset_index(drop=True)
                    self.logger.debug(f"清理后的数据行数: {len(df)}")
                    
                    if not df.empty:
                        self.logger.debug("\n清理后的数据:")
                        for index, row in df.iterrows():
                            self.logger.debug(f"  行 {index+1}: {list(row)}")
            
            self.logger.debug("=== 读取客户数据完成 ===")
            return df
        except (FileNotFoundError, PermissionError, ValueError, KeyError, OSError) as e:
            self.logger.error(f"获取客户数据时出错: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def get_customer_by_id(self, customer_id):
        df = self.read_sheet("客户信息")
        if df.empty:
            return None
        result = df[df['客户编号'] == customer_id]
        return result.iloc[0] if not result.empty else None

    def get_customer_by_name(self, name):
        df = self.read_sheet("客户信息")
        if df.empty:
            return None
        result = df[df['客户名称'] == name]
        return result.iloc[0] if not result.empty else None

    def add_customer(self, customer_data):
        self.append_to_sheet("客户信息", customer_data)
        return {'success': True}

    def update_customer(self, customer_id, new_data):
        df = self.read_sheet("客户信息")
        if df.empty:
            return False

        idx = df[df['客户编号'] == customer_id].index
        if len(idx) > 0:
            for col, value in new_data.items():
                df.at[idx[0], col] = value
            self.write_sheet("客户信息", df)
            return True
        return False

    def delete_customer(self, customer_id):
        df = self.read_sheet("客户信息")
        if df.empty:
            return False

        df = df[df['客户编号'] != customer_id]
        self.write_sheet("客户信息", df)
        return True

    @staticmethod
    def calculate_customer_level(total_purchases):
        """统一计算客户等级
        
        Args:
            total_purchases: 累计消费金额
            
        Returns:
            客户等级字符串
        """
        if total_purchases >= 10000:
            return "VIP客户"
        elif total_purchases >= 5000:
            return "高级客户"
        elif total_purchases >= 2000:
            return "中级客户"
        else:
            return "普通客户"
    
    def update_customer_after_sale(self, customer_name, amount, sale_date):
        """销售后更新客户信息"""
        df = self.read_sheet("客户信息")
        if df.empty:
            # 如果客户表为空，创建新客户
            self._create_new_customer(customer_name, amount, sale_date)
            return

        # 查找客户
        customer_row = df[df['客户名称'] == customer_name]
        if customer_row.empty:
            # 创建新客户
            self._create_new_customer(customer_name, amount, sale_date)
        else:
            # 更新现有客户
            idx = customer_row.index[0]
            current_purchases = float(df.at[idx, '累计消费']) if pd.notna(df.at[idx, '累计消费']) else 0.0
            current_orders = int(df.at[idx, '订单数']) if pd.notna(df.at[idx, '订单数']) else 0
            
            # 更新客户信息
            new_purchases = current_purchases + amount
            new_orders = current_orders + 1
            
            # 使用统一函数计算客户等级
            customer_level = self.calculate_customer_level(new_purchases)
            
            # 完全重新创建列来确保数据类型正确
            purchases_list = df['累计消费'].tolist()
            purchases_list[idx] = float(new_purchases)
            df['累计消费'] = pd.Series(purchases_list, dtype=float)
            
            orders_list = df['订单数'].tolist()
            orders_list[idx] = int(new_orders)
            df['订单数'] = pd.Series(orders_list, dtype=int)
            
            df.at[idx, '最后购买日期'] = sale_date
            df.at[idx, '客户等级'] = customer_level
            
            self.write_sheet("客户信息", df)

    def _create_new_customer(self, customer_name, amount, sale_date):
        """创建新客户"""
        # 生成客户编号
        customer_id = self.generate_id("K", "客户信息", "客户编号")
        
        # 使用统一函数确定客户等级
        customer_level = self.calculate_customer_level(amount)
        
        # 创建客户数据
        customer_data = [
            customer_id, customer_name, "", "", "",
            amount, 1, sale_date, customer_level, "",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
        
        self.append_to_sheet("客户信息", customer_data)

    # 销售记录查询与修改
    def get_sale_by_id(self, sale_id):
        df = self.read_sheet("销售记录")
        if df.empty:
            return None
        result = df[df['销售编号'] == sale_id]
        return result.iloc[0] if not result.empty else None

    def query_sales(self, start_date=None, end_date=None, customer_name=None,
                    com_id=None, com_name=None, include_voided=False):
        """多条件查询销售记录
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            customer_name: 客户名称
            com_id: 商品编号
            com_name: 商品名称
            include_voided: 是否包含作废记录，默认为 False
            
        Returns:
            查询结果 DataFrame
        """
        df = self.read_sheet("销售记录")
        if df.empty:
            return df

        # 统一过滤作废记录
        if not include_voided:
            if '是否作废' in df.columns:
                df = df[df['是否作废'] != True]

        # 按日期范围过滤
        if start_date:
            df = df[df['销售日期'] >= start_date]
        if end_date:
            df = df[df['销售日期'] <= end_date]

        # 按客户过滤
        if customer_name:
            df = df[df['客户名称'] == customer_name]

        # 按商品过滤
        if com_id:
            df = df[df['商品编号'] == com_id]
        if com_name:
            df = df[df['商品名称'] == com_name]

        return df

    def update_sale(self, sale_id, new_data, rollback_stock=False):
        """修改销售记录，可选回滚库存"""
        df = self.read_sheet("销售记录")
        if df.empty:
            return False

        idx = df[df['销售编号'] == sale_id].index
        if len(idx) == 0:
            return False

        old_sale = df.iloc[idx[0]]

        # 如果需要回滚库存
        if rollback_stock:
            # 恢复原库存
            commodity = self.get_commodity_by_id(old_sale['商品编号'])
            if commodity is not None:
                old_qty_jin = convert_to_jin(old_sale['销售数量'], old_sale.get('销售单位', '克'))
                current_stock = float(commodity['当前库存'])
                self.update_commodity(old_sale['商品编号'], {'当前库存': current_stock + old_qty_jin})

            # 如果有新数据，扣减新库存
            if '销售数量' in new_data or '销售单位' in new_data:
                new_qty = new_data.get('销售数量', old_sale['销售数量'])
                new_unit = new_data.get('销售单位', old_sale.get('销售单位', '克'))
                new_qty_jin = convert_to_jin(new_qty, new_unit)
                commodity = self.get_commodity_by_id(old_sale['商品编号'])
                if commodity is not None:
                    current_stock = float(commodity['当前库存'])
                    self.update_commodity(old_sale['商品编号'], {'当前库存': current_stock - new_qty_jin})

        # 更新销售记录
        for col, value in new_data.items():
            df.at[idx[0], col] = value

        self.write_sheet("销售记录", df)
        return True

    def flush(self):
        """批量写入所有待处理的sheet到Excel文件"""
        if not self._dirty or not self._pending_writes:
            return

        backup_filename = None
        try:
            if os.path.exists(self.filename):
                backup_filename = self.filename + ".backup"
                shutil.copy2(self.filename, backup_filename)

            wb = load_workbook(self.filename) if os.path.exists(self.filename) else Workbook()
            if not wb.sheetnames:
                wb.active.title = "Sheet"

            for sheet_name, data_to_write in self._pending_writes.items():
                existing_columns = None
                sheet_exists = sheet_name in wb.sheetnames

                if sheet_exists:
                    ws = wb[sheet_name]
                    if ws.max_row > 0:
                        existing_columns = [cell.value for cell in ws[1]]

                new_columns = data_to_write.columns.tolist()

                if sheet_exists and existing_columns == new_columns:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                        for cell in row:
                            cell.value = None
                    ws.delete_rows(2, ws.max_row - 1)
                    ws = wb[sheet_name]
                else:
                    if sheet_exists:
                        wb.remove(wb[sheet_name])
                    ws = wb.create_sheet(sheet_name)
                    if new_columns:
                        ws.append(new_columns)

                if not data_to_write.empty:
                    row_count = 0
                    for row in dataframe_to_rows(data_to_write, index=False, header=False):
                        ws.append(row)
                        row_count += 1

            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb["Sheet"])

            wb.save(self.filename)

            if backup_filename and os.path.exists(backup_filename):
                os.remove(backup_filename)

            self._pending_writes.clear()
            self._dirty = False
        except (PermissionError, OSError) as e:
            self.logger.error(f"批量写入Excel文件出错: {e}")
            if backup_filename and os.path.exists(backup_filename):
                try:
                    shutil.copy2(backup_filename, self.filename)
                    os.remove(backup_filename)
                    self.logger.info("写入失败，已回滚到备份文件")
                except Exception as rollback_error:
                    self.logger.error(f"回滚失败: {rollback_error}")
            raise

    def auto_flush(self):
        """返回一个上下文管理器，退出时自动调用 flush()"""
        return _AutoFlushContext(self)

    def __del__(self):
        try:
            self.flush()
        except Exception:
            pass


    def void_sale(self, sale_id):
        """作废销售记录，回滚库存"""
        df = self.read_sheet("销售记录")
        if df.empty:
            return False

        idx = df[df['销售编号'] == sale_id].index
        if len(idx) == 0:
            return False

        sale = df.iloc[idx[0]]

        # 回滚库存
        commodity = self.get_commodity_by_id(sale['商品编号'])
        if commodity is not None:
            qty_jin = convert_to_jin(sale['销售数量'], sale.get('销售单位', '克'))
            current_stock = float(commodity['当前库存'])
            self.update_commodity(sale['商品编号'], {'当前库存': current_stock + qty_jin})

        # 标记为作废
        df.at[idx[0], '是否作废'] = True
        self.write_sheet("销售记录", df)
        return True

    def delete_sale(self, sale_id):
        """删除销售记录并完整回退相关数据（库存加回、客户累计消费/订单数/等级/最后购买日期回退）

        Args:
            sale_id: 销售编号

        Returns:
            bool: 是否删除成功
        """
        df = self.read_sheet("销售记录")
        if df.empty:
            return False

        idx_list = df[df['销售编号'] == sale_id].index
        if len(idx_list) == 0:
            return False
        idx = idx_list[0]
        sale = df.iloc[idx]

        # 1. 回滚库存：当前库存 + 销售数量(转换为斤)
        commodity = self.get_commodity_by_id(sale['商品编号'])
        if commodity is not None:
            qty_jin = convert_to_jin(sale['销售数量'], sale.get('销售单位', '克'))
            current_stock = float(commodity['当前库存'])
            self.update_commodity(sale['商品编号'], {'当前库存': current_stock + qty_jin})

        # 2. 回退客户信息：累计消费减少、订单数减少、客户等级重新计算、最后购买日期更新
        customer_name = sale.get('客户名称', '')
        received = sale.get('实收金额', 0)
        if customer_name and pd.notna(received):
            self._revert_customer_after_sale(customer_name, sale_id, received)

        # 3. 从销售记录表中删除该行
        new_df = df.drop(index=idx)
        self.write_sheet("销售记录", new_df)
        return True

    def _revert_customer_after_sale(self, customer_name, sale_id, amount):
        """销售记录删除后回退客户信息"""
        df = self.read_sheet("客户信息")
        if df.empty:
            return

        customer_row = df[df['客户名称'] == customer_name]
        if customer_row.empty:
            return

        idx = customer_row.index[0]
        current_purchases = float(df.at[idx, '累计消费']) if pd.notna(df.at[idx, '累计消费']) else 0.0
        current_orders = int(df.at[idx, '订单数']) if pd.notna(df.at[idx, '订单数']) else 0

        new_purchases = max(0.0, current_purchases - float(amount))
        new_orders = max(0, current_orders - 1)

        # 重新计算客户等级
        customer_level = self.calculate_customer_level(new_purchases)

        # 完全重新创建列来确保数据类型正确
        purchases_list = df['累计消费'].tolist()
        purchases_list[idx] = float(new_purchases)
        df['累计消费'] = pd.Series(purchases_list, dtype=float)

        orders_list = df['订单数'].tolist()
        orders_list[idx] = int(new_orders)
        df['订单数'] = pd.Series(orders_list, dtype=int)

        # 最后购买日期：取该客户剩余（未删除）销售记录中最新的日期
        last_date = ""
        sales_df = self.read_sheet("销售记录")
        if not sales_df.empty and '客户名称' in sales_df.columns:
            remaining = sales_df[(sales_df['客户名称'] == customer_name) & (sales_df['销售编号'] != sale_id)]
            if not remaining.empty and '销售日期' in remaining.columns:
                last_date = remaining['销售日期'].max()
        df.at[idx, '最后购买日期'] = last_date
        df.at[idx, '客户等级'] = customer_level

        self.write_sheet("客户信息", df)


class _AutoFlushContext:
    def __init__(self, manager):
        self._manager = manager

    def __enter__(self):
        return self._manager

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._manager.flush()
        return False