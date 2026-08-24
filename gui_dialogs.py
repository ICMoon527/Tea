import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from logger import get_logger
from styles import Styles

_logger = get_logger()


def select_product_dialog(parent, target_var):
    """商品选择弹窗 - 双击选择商品

    Args:
        parent: TeaInventoryGUI 实例
        target_var: 要填充的StringVar变量
    """
    df = parent.system.excel_manager.get_all_commodities()

    top = parent._create_toplevel_with_size("select_product", "large")
    top.title("选择商品")
    top.configure(bg=Styles.BACKGROUND_COLOR)
    top.resizable(True, True)

    title_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    title_frame.pack(pady=Styles.PADY_MEDIUM)

    tk.Label(
        title_frame,
        text="选择商品（双击选择）",
        font=Styles.SUB_HEADER_FONT,
        bg=Styles.BACKGROUND_COLOR,
        fg=Styles.HEADER_COLOR
    ).pack()

    if df.empty:
        tk.Label(top, text="暂无商品数据", font=Styles.LABEL_FONT, bg=Styles.BACKGROUND_COLOR, fg=Styles.TEXT_COLOR).pack(pady=Styles.PADY_LARGE)

        btn_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
        btn_frame.pack(pady=Styles.PADY_MEDIUM)

        btn_add = tk.Button(btn_frame, text="手动添加商品", font=Styles.BUTTON_FONT,
                           width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                           command=lambda: [top.destroy(), parent.add_product_gui()],
                           bg=Styles.PRIMARY_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
        btn_add.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)
        btn_add.bind("<Enter>", lambda e, b=btn_add: b.config(bg=Styles.BUTTON_HOVER_COLOR))
        btn_add.bind("<Leave>", lambda e, b=btn_add: b.config(bg=Styles.PRIMARY_COLOR))

        btn_cancel = tk.Button(btn_frame, text="取消", font=Styles.BUTTON_FONT,
                               width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                               command=top.destroy,
                               bg=Styles.ERROR_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
        btn_cancel.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)
        return

    list_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    list_frame.pack(pady=Styles.PADY_MEDIUM, padx=Styles.PADX_MEDIUM, fill=tk.BOTH, expand=True)

    tree = ttk.Treeview(list_frame, style="Treeview", show="headings")
    tree["columns"] = ("商品编号", "商品名称", "茶类", "品种", "当前库存", "零售价")

    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=140, anchor=tk.CENTER)

    scrollbar_y = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscrollcommand=scrollbar_y.set)

    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    for _, row in df.iterrows():
        tree.insert("", tk.END, values=(
            row['商品编号'],
            row['商品名称'],
            row['茶类'],
            row['品种'],
            row['当前库存'],
            row['零售价']
        ))

    def on_double_click(event):
        selection = tree.selection()
        if selection:
            item = tree.item(selection[0])
            com_id = item['values'][0]
            target_var.set(com_id)
            top.destroy()

    tree.bind('<Double-1>', on_double_click)

    btn_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    btn_frame.pack(pady=Styles.PADY_MEDIUM)

    btn_add = tk.Button(btn_frame, text="手动添加商品", font=Styles.BUTTON_FONT,
                       width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                       command=lambda: [top.destroy(), parent.add_product_gui()],
                       bg=Styles.SECONDARY_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
    btn_add.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)

    btn_cancel = tk.Button(btn_frame, text="取消", font=Styles.BUTTON_FONT,
                           width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                           command=top.destroy,
                           bg=Styles.ERROR_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
    btn_cancel.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)


def select_supplier_dialog(parent, target_var):
    """供应商选择弹窗 - 双击选择供应商

    Args:
        parent: TeaInventoryGUI 实例
        target_var: 要填充的StringVar变量
    """
    df = parent.system.excel_manager.get_all_suppliers()

    top = parent._create_toplevel_with_size("select_supplier", "large")
    top.title("选择供应商")
    top.configure(bg=Styles.BACKGROUND_COLOR)
    top.resizable(True, True)

    title_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    title_frame.pack(pady=Styles.PADY_MEDIUM)

    tk.Label(
        title_frame,
        text="选择供应商（双击选择）",
        font=Styles.SUB_HEADER_FONT,
        bg=Styles.BACKGROUND_COLOR,
        fg=Styles.HEADER_COLOR
    ).pack()

    if df.empty:
        tk.Label(top, text="暂无供应商数据", font=Styles.LABEL_FONT, bg=Styles.BACKGROUND_COLOR, fg=Styles.TEXT_COLOR).pack(pady=Styles.PADY_LARGE)

        btn_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
        btn_frame.pack(pady=Styles.PADY_MEDIUM)

        btn_add = tk.Button(btn_frame, text="手动添加供应商", font=Styles.BUTTON_FONT,
                           width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                           command=lambda: [top.destroy(), parent.add_supplier_gui()],
                           bg=Styles.PRIMARY_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
        btn_add.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)
        btn_add.bind("<Enter>", lambda e, b=btn_add: b.config(bg=Styles.BUTTON_HOVER_COLOR))
        btn_add.bind("<Leave>", lambda e, b=btn_add: b.config(bg=Styles.PRIMARY_COLOR))

        btn_cancel = tk.Button(btn_frame, text="取消", font=Styles.BUTTON_FONT,
                               width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                               command=top.destroy,
                               bg=Styles.ERROR_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
        btn_cancel.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)
        return

    list_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    list_frame.pack(pady=Styles.PADY_MEDIUM, padx=Styles.PADX_MEDIUM, fill=tk.BOTH, expand=True)

    tree = ttk.Treeview(list_frame, style="Treeview", show="headings")
    tree["columns"] = ("供应商编号", "供应商名称", "联系电话", "地址", "累计交易金额")

    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=160, anchor=tk.CENTER)

    scrollbar_y = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscrollcommand=scrollbar_y.set)

    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    for _, row in df.iterrows():
        tree.insert("", tk.END, values=(
            row.get('供应商编号', ''),
            row.get('供应商名称', ''),
            '' if pd.isna(row.get('联系电话')) else row.get('联系电话', ''),
            '' if pd.isna(row.get('地址')) else row.get('地址', ''),
            row.get('累计交易金额', 0) if pd.notna(row.get('累计交易金额')) else 0
        ))

    def on_double_click(event):
        selection = tree.selection()
        if selection:
            item = tree.item(selection[0])
            supplier_name = item['values'][1]
            target_var.set(supplier_name)
            top.destroy()

    tree.bind('<Double-1>', on_double_click)

    btn_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    btn_frame.pack(pady=Styles.PADY_MEDIUM)

    btn_add = tk.Button(btn_frame, text="手动添加供应商", font=Styles.BUTTON_FONT,
                       width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                       command=lambda: [top.destroy(), parent.add_supplier_gui()],
                       bg=Styles.SECONDARY_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
    btn_add.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)

    btn_cancel = tk.Button(btn_frame, text="取消", font=Styles.BUTTON_FONT,
                           width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                           command=top.destroy,
                           bg=Styles.ERROR_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
    btn_cancel.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)


def select_customer_dialog(parent, target_var):
    """客户选择弹窗 - 双击选择客户

    Args:
        parent: TeaInventoryGUI 实例
        target_var: 要填充的StringVar变量
    """
    df = parent.system.excel_manager.get_all_customers()

    top = parent._create_toplevel_with_size("select_customer", "large")
    top.title("选择客户")
    top.configure(bg=Styles.BACKGROUND_COLOR)
    top.resizable(True, True)

    title_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    title_frame.pack(pady=Styles.PADY_MEDIUM)

    tk.Label(
        title_frame,
        text="选择客户（双击选择）",
        font=Styles.SUB_HEADER_FONT,
        bg=Styles.BACKGROUND_COLOR,
        fg=Styles.HEADER_COLOR
    ).pack()

    if df.empty:
        tk.Label(top, text="暂无客户数据", font=Styles.LABEL_FONT, bg=Styles.BACKGROUND_COLOR, fg=Styles.TEXT_COLOR).pack(pady=Styles.PADY_LARGE)

        btn_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
        btn_frame.pack(pady=Styles.PADY_MEDIUM)

        btn_add = tk.Button(btn_frame, text="手动添加客户", font=Styles.BUTTON_FONT,
                           width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                           command=lambda: [top.destroy(), parent.add_customer_gui()],
                           bg=Styles.PRIMARY_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
        btn_add.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)
        btn_add.bind("<Enter>", lambda e, b=btn_add: b.config(bg=Styles.BUTTON_HOVER_COLOR))
        btn_add.bind("<Leave>", lambda e, b=btn_add: b.config(bg=Styles.PRIMARY_COLOR))

        btn_cancel = tk.Button(btn_frame, text="取消", font=Styles.BUTTON_FONT,
                               width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                               command=top.destroy,
                               bg=Styles.ERROR_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
        btn_cancel.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)
        return

    # 模糊搜索框（类似商品列表的模糊搜索）
    search_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    search_frame.pack(pady=(Styles.PADY_SMALL, 0))
    search_var = tk.StringVar()
    tk.Label(search_frame, text="搜索客户: ", font=Styles.LABEL_FONT,
             bg=Styles.BACKGROUND_COLOR, fg=Styles.TEXT_COLOR).pack(side=tk.LEFT)
    search_entry = tk.Entry(search_frame, textvariable=search_var,
                            font=Styles.TEXT_FONT, width=30)
    search_entry.pack(side=tk.LEFT, padx=(0, 5))
    tk.Label(search_frame, text="支持名称模糊搜索", font=Styles.TEXT_FONT,
             bg=Styles.BACKGROUND_COLOR, fg=Styles.TEXT_COLOR).pack(side=tk.LEFT)

    list_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    list_frame.pack(pady=Styles.PADY_MEDIUM, padx=Styles.PADX_MEDIUM, fill=tk.BOTH, expand=True)

    tree = ttk.Treeview(list_frame, style="Treeview", show="headings")
    tree["columns"] = ("客户编号", "客户名称", "联系电话", "地址", "累计消费")

    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=160, anchor=tk.CENTER)

    scrollbar_y = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscrollcommand=scrollbar_y.set)

    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    def _populate(filtered_df):
        """清空并按筛选结果填充客户列表"""
        for item in tree.get_children():
            tree.delete(item)
        for _, row in filtered_df.iterrows():
            tree.insert("", tk.END, values=(
                row.get('客户编号', ''),
                row.get('客户名称', ''),
                '' if pd.isna(row.get('联系电话')) else row.get('联系电话', ''),
                '' if pd.isna(row.get('地址')) else row.get('地址', ''),
                row.get('累计消费', 0) if pd.notna(row.get('累计消费')) else 0
            ))

    def _filter_customers(*_):
        """根据关键词模糊筛选客户"""
        keyword = search_var.get().strip()
        if not keyword:
            _populate(df)
            return
        mask = df['客户名称'].str.contains(keyword, case=False, na=False)
        _populate(df[mask])

    _populate(df)
    search_entry.bind('<KeyRelease>', _filter_customers)
    search_entry.bind('<Return>', _filter_customers)

    def on_double_click(event):
        selection = tree.selection()
        if selection:
            item = tree.item(selection[0])
            customer_name = item['values'][1]
            target_var.set(customer_name)
            top.destroy()

    tree.bind('<Double-1>', on_double_click)

    btn_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    btn_frame.pack(pady=Styles.PADY_MEDIUM)

    btn_add = tk.Button(btn_frame, text="手动添加客户", font=Styles.BUTTON_FONT,
                       width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                       command=lambda: [top.destroy(), parent.add_customer_gui()],
                       bg=Styles.SECONDARY_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
    btn_add.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)

    btn_cancel = tk.Button(btn_frame, text="取消", font=Styles.BUTTON_FONT,
                           width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT,
                           command=top.destroy,
                           bg=Styles.ERROR_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
    btn_cancel.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)


def show_dataframe_window(parent, df, title):
    """显示DataFrame的窗口（带分页和搜索功能）

    Args:
        parent: TeaInventoryGUI 实例
        df: 要显示的DataFrame，或 Excel 文件路径(str)以启用真正按需分页
        title: 窗口标题
    """
    window_id = "dataframe_" + title.replace(" ", "_").replace("（", "").replace("）", "").replace("-", "_")
    top = parent._create_toplevel_with_size(window_id, "large")
    top.title(title)
    top.configure(bg=Styles.BACKGROUND_COLOR)

    PAGE_SIZE = 100
    current_page_var = tk.IntVar(value=1)

    _file_path = None
    if isinstance(df, str):
        _file_path = df
        from openpyxl import load_workbook
        wb = load_workbook(_file_path, read_only=True)
        ws = wb.active
        total_rows = ws.max_row - 1
        wb.close()
        sample_page = pd.read_excel(_file_path, engine='openpyxl', nrows=PAGE_SIZE)
        sample_page.columns = [str(c) for c in sample_page.columns]
        columns = list(sample_page.columns)
        df = sample_page
        _original_df = None
    else:
        total_rows = len(df)
        df.columns = [str(col) for col in df.columns]
        columns = list(df.columns)

    _logger.debug(f"show_dataframe_window: 标题={title}, 行数={total_rows}, 列数={len(columns)}, 是否为空={total_rows == 0}")
    _logger.debug(f"列名: {columns}")

    if total_rows == 0:
        frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
        frame.pack(expand=True, fill=tk.BOTH)

        tk.Label(
            frame,
            text="暂无数据",
            font=Styles.SUB_HEADER_FONT,
            bg=Styles.BACKGROUND_COLOR,
            fg=Styles.TEXT_COLOR
        ).pack(expand=True)

        btn_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
        btn_frame.pack(pady=Styles.PADY_MEDIUM)

        tk.Button(
            btn_frame,
            text="关闭",
            font=Styles.BUTTON_FONT,
            width=Styles.BUTTON_WIDTH,
            height=Styles.BUTTON_HEIGHT,
            command=top.destroy,
            bg=Styles.ERROR_COLOR,
            fg="white",
            relief=tk.FLAT,
            padx=10,
            pady=5
        ).pack()
        return

    def _apply_transforms(frame_df):
        if '销售单位' in frame_df.columns and '销售数量' in frame_df.columns:
            for idx, row in frame_df.iterrows():
                if row['销售单位'] == '克' and pd.notna(row['销售数量']):
                    frame_df.at[idx, '销售数量'] = row['销售数量'] / 500
                    frame_df.at[idx, '销售单位'] = '斤'
        if '实收金额' in frame_df.columns:
            for idx, row in frame_df.iterrows():
                val = row['实收金额']
                if pd.notna(val):
                    try:
                        frame_df.at[idx, '实收金额'] = round(float(val), 1)
                    except:
                        pass
        return frame_df

    if not _file_path:
        df = _apply_transforms(df.copy())
        _original_df = df
    else:
        df = _apply_transforms(df.copy())

    _search_filtered_df = None

    title_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    title_frame.pack(pady=Styles.PADY_SMALL, fill=tk.X)

    tk.Label(
        title_frame,
        text=title,
        font=Styles.LABEL_FONT,
        bg=Styles.BACKGROUND_COLOR,
        fg=Styles.HEADER_COLOR
    ).pack(padx=Styles.PADX_MEDIUM, anchor=tk.W)

    search_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    search_frame.pack(pady=(5, 0), fill=tk.X)

    search_var = tk.StringVar()
    search_entry = tk.Entry(search_frame, textvariable=search_var, font=Styles.TEXT_FONT, width=30)
    search_entry.pack(padx=Styles.PADX_MEDIUM, side=tk.LEFT)

    PLACEHOLDER = "输入关键字搜索..."
    search_entry.insert(0, PLACEHOLDER)
    search_entry.config(fg="#999999")

    def on_search_focus_in(event):
        if search_entry.get() == PLACEHOLDER:
            search_entry.delete(0, tk.END)
            search_entry.config(fg="black")

    def on_search_focus_out(event):
        if not search_entry.get().strip():
            search_entry.delete(0, tk.END)
            search_entry.insert(0, PLACEHOLDER)
            search_entry.config(fg="#999999")

    search_entry.bind("<FocusIn>", on_search_focus_in)
    search_entry.bind("<FocusOut>", on_search_focus_out)

    table_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    table_frame.pack(padx=Styles.PADX_MEDIUM, pady=Styles.PADY_SMALL, fill=tk.BOTH, expand=True)

    tree = ttk.Treeview(table_frame, style="Treeview")
    tree["columns"] = columns
    tree["show"] = "headings"

    column_widths = {
        '茶类': 100,
        '品种': 120,
        '商品编号': 100,
        '商品名称': 150,
        '销售数量(斤)': 100,
        '实收金额': 100,
        '销售成本': 100,
        '利润': 100,
        '利润率(%)': 100,
        '日期': 120,
        '日': 100,
        '周': 120,
        '月': 100,
        '供应商编号': 100,
        '供应商名称': 150,
        '联系人': 100,
        '联系电话': 120,
        '客户编号': 100,
        '客户名称': 150,
        '累计消费': 100,
        '客户等级': 100,
        '销售编号': 120,
        '进货编号': 120,
        '销售日期': 120,
        '进货日期': 120,
        '销售单位': 80,
        '进货单位': 80,
        '公司': 120,
        '产区': 100,
        '规格': 80,
        '成本价': 80,
        '零售价': 80,
        '生产日期': 100,
        '保质期(月)': 90,
        '当前库存': 90,
        '品质特征': 150,
        '年份': 60,
        '等级': 60,
        '单位': 60,
        '销售数量': 90,
        '进货数量': 90,
        '单价': 80,
        '进货单价': 90,
        '应收金额': 100,
        '供应商': 120,
        '地址': 150,
        '备注': 120,
        '电子邮箱': 150,
        '订单数': 80,
        '最后购买日期': 120,
        '创建日期': 120,
        '是否作废': 80,
        '预警级别': 90,
        '过期日期': 120,
        '剩余天数': 90
    }

    for col in columns:
        tree.heading(col, text=str(col))
        width = column_widths.get(str(col), 120)
        tree.column(col, width=width, anchor=tk.CENTER)

    scrollbar_y = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
    scrollbar_x = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=tree.xview)

    tree.configure(
        yscrollcommand=scrollbar_y.set,
        xscrollcommand=scrollbar_x.set
    )

    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
    scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    _columns_fitted = False

    def _auto_fit_columns(data_df, max_width=300, min_width=60):
        for col_idx, col_name in enumerate(data_df.columns):
            header_width = len(str(col_name)) * 10

            col_data = data_df[col_name].astype(str)
            if len(col_data) > 0:
                max_content_len = col_data.str.len().max()
            else:
                max_content_len = 0
            content_width = max_content_len * 9

            width = max(header_width, content_width, min_width)
            width = min(width, max_width)

            tree.column(f"#{col_idx + 1}", width=int(width))

    def _auto_fit_single_column(data_df, col_name, col_idx, max_width=300, min_width=60):
        header_width = len(str(col_name)) * 10

        col_data = data_df[col_name].astype(str)
        if len(col_data) > 0:
            max_content_len = col_data.str.len().max()
        else:
            max_content_len = 0
        content_width = max_content_len * 9

        width = max(header_width, content_width, min_width)
        width = min(width, max_width)

        tree.column(f"#{col_idx + 1}", width=int(width))

    def _load_page_data(page_num):
        if _search_filtered_df is not None:
            working = _search_filtered_df
            start_idx = (page_num - 1) * PAGE_SIZE
            end_idx = min(start_idx + PAGE_SIZE, len(working))
            return working.iloc[start_idx:end_idx], len(working)

        if _file_path:
            if page_num == 1:
                page_df = pd.read_excel(_file_path, engine='openpyxl', nrows=PAGE_SIZE)
            else:
                start_row = (page_num - 1) * PAGE_SIZE + 1
                page_df = pd.read_excel(_file_path, engine='openpyxl',
                                        skiprows=range(1, start_row),
                                        header=0,
                                        nrows=PAGE_SIZE)
            page_df.columns = [str(c) for c in page_df.columns]
            _apply_transforms(page_df)
            return page_df, total_rows

        start_idx = (page_num - 1) * PAGE_SIZE
        end_idx = min(start_idx + PAGE_SIZE, total_rows)
        return _original_df.iloc[start_idx:end_idx], total_rows

    def refresh_page():
        nonlocal _columns_fitted
        for item in tree.get_children():
            tree.delete(item)

        page_num = current_page_var.get()
        page_df, working_total = _load_page_data(page_num)
        working_pages = max(1, (working_total + PAGE_SIZE - 1) // PAGE_SIZE)

        for _, row in page_df.iterrows():
            values = []
            for col in columns:
                val = row[col]
                if pd.isna(val):
                    values.append("")
                elif col == "实收金额":
                    try:
                        num_val = float(val)
                        values.append(f"{num_val:.1f}")
                    except:
                        values.append(str(val))
                else:
                    values.append(str(val))
            tree.insert("", tk.END, values=values)

        if not _columns_fitted:
            if _original_df is not None:
                _auto_fit_columns(_original_df)
            else:
                _auto_fit_columns(page_df)
            _columns_fitted = True

        page_info_label.config(text=f"第 {page_num} / {working_pages} 页 (共 {working_total} 条记录)")

        prev_btn.config(state=tk.NORMAL if page_num > 1 else tk.DISABLED)
        next_btn.config(state=tk.NORMAL if page_num < working_pages else tk.DISABLED)
        first_btn.config(state=tk.NORMAL if page_num > 1 else tk.DISABLED)
        last_btn.config(state=tk.NORMAL if page_num < working_pages else tk.DISABLED)

    def on_search_keyrelease(event):
        search_text = search_var.get().strip()
        if search_text == PLACEHOLDER:
            search_text = ""

        nonlocal _search_filtered_df, _columns_fitted
        if not search_text:
            _search_filtered_df = None
        else:
            if _file_path:
                full = pd.read_excel(_file_path, engine='openpyxl')
                full.columns = [str(c) for c in full.columns]
                full = _apply_transforms(full.copy())
            else:
                full = _original_df

            mask = pd.Series([False] * len(full))
            for col in columns:
                col_str = full[col].astype(str)
                mask = mask | col_str.str.contains(search_text, case=False, na=False)
            _search_filtered_df = full[mask].copy()

        _columns_fitted = False
        current_page_var.set(1)
        refresh_page()

    search_entry.bind("<KeyRelease>", on_search_keyrelease)

    def go_first():
        current_page_var.set(1)
        refresh_page()

    def go_prev():
        current = current_page_var.get()
        if current > 1:
            current_page_var.set(current - 1)
            refresh_page()

    def go_next():
        current = current_page_var.get()
        working_total = len(_search_filtered_df) if _search_filtered_df is not None else total_rows
        working_pages = max(1, (working_total + PAGE_SIZE - 1) // PAGE_SIZE)
        if current < working_pages:
            current_page_var.set(current + 1)
            refresh_page()

    def go_last():
        working_total = len(_search_filtered_df) if _search_filtered_df is not None else total_rows
        working_pages = max(1, (working_total + PAGE_SIZE - 1) // PAGE_SIZE)
        current_page_var.set(working_pages)
        refresh_page()

    pagination_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    pagination_frame.pack(pady=Styles.PADY_SMALL)

    first_btn = tk.Button(pagination_frame, text="首页", font=Styles.TEXT_FONT,
                       command=go_first, bg=Styles.PRIMARY_COLOR, fg="white",
                       relief=tk.FLAT, padx=10, pady=3)
    first_btn.pack(side=tk.LEFT, padx=5)

    prev_btn = tk.Button(pagination_frame, text="上一页", font=Styles.TEXT_FONT,
                      command=go_prev, bg=Styles.PRIMARY_COLOR, fg="white",
                      relief=tk.FLAT, padx=10, pady=3)
    prev_btn.pack(side=tk.LEFT, padx=5)

    page_info_label = tk.Label(pagination_frame, text="",
                              font=Styles.LABEL_FONT,
                              bg=Styles.BACKGROUND_COLOR,
                              fg=Styles.TEXT_COLOR)
    page_info_label.pack(side=tk.LEFT, padx=20)

    next_btn = tk.Button(pagination_frame, text="下一页", font=Styles.TEXT_FONT,
                      command=go_next, bg=Styles.PRIMARY_COLOR, fg="white",
                      relief=tk.FLAT, padx=10, pady=3)
    next_btn.pack(side=tk.LEFT, padx=5)

    last_btn = tk.Button(pagination_frame, text="末页", font=Styles.TEXT_FONT,
                       command=go_last, bg=Styles.PRIMARY_COLOR, fg="white",
                       relief=tk.FLAT, padx=10, pady=3)
    last_btn.pack(side=tk.LEFT, padx=5)

    refresh_page()

    def show_detail(e):
        selected = tree.selection()
        if not selected:
            return

        item = tree.item(selected[0])
        values = item['values']

        if '商品编号' in columns:
            com_id_idx = columns.index('商品编号')
            com_id = values[com_id_idx]

            commodity = parent.system.excel_manager.get_commodity_by_id(com_id)
            if commodity is not None:
                info_top = parent._create_toplevel_with_size("dataframe_product_detail", "medium", parent=top)
                info_top.title("商品详情")
                info_top.configure(bg=Styles.BACKGROUND_COLOR)
                info_top.resizable(True, True)

                frame = tk.Frame(info_top, bg=Styles.BACKGROUND_COLOR)
                frame.pack(pady=Styles.PADY_LARGE, padx=Styles.PADX_LARGE)

                for i, (key, value) in enumerate(commodity.items()):
                    row_frame = tk.Frame(frame, bg=Styles.BACKGROUND_COLOR)
                    row_frame.pack(fill=tk.X, pady=5)

                    tk.Label(row_frame, text=f"{key}:",
                            font=Styles.LABEL_FONT,
                            bg=Styles.BACKGROUND_COLOR,
                            fg=Styles.HEADER_COLOR,
                            width=15, anchor=tk.W).pack(side=tk.LEFT)

                    tk.Label(row_frame, text=str(value) if pd.notna(value) else "",
                            font=Styles.TEXT_FONT,
                            bg=Styles.BACKGROUND_COLOR,
                            fg=Styles.TEXT_COLOR,
                            anchor=tk.W).pack(side=tk.LEFT, fill=tk.X, expand=True)

                tk.Button(info_top, text="关闭",
                         font=Styles.BUTTON_FONT,
                         width=Styles.BUTTON_WIDTH,
                         height=Styles.BUTTON_HEIGHT,
                         command=info_top.destroy,
                         bg=Styles.ERROR_COLOR,
                         fg="white",
                         relief=tk.FLAT,
                         padx=10,
                         pady=5).pack(pady=Styles.PADY_MEDIUM)

        elif '供应商编号' in columns:
            supplier_id_idx = columns.index('供应商编号')
            supplier_id = values[supplier_id_idx]

            df_supplier = parent.system.excel_manager.get_all_suppliers()
            if not df_supplier.empty:
                supplier = df_supplier[df_supplier['供应商编号'] == supplier_id]
                if not supplier.empty:
                    supplier_row = supplier.iloc[0]

                    edit_top = parent._create_toplevel_with_size("dataframe_edit_supplier", "medium", parent=top)
                    edit_top.title("修改供应商信息")
                    edit_top.configure(bg=Styles.BACKGROUND_COLOR)
                    edit_top.resizable(True, True)

                    edit_title_frame = tk.Frame(edit_top, bg=Styles.BACKGROUND_COLOR)
                    edit_title_frame.pack(pady=Styles.PADY_MEDIUM)

                    tk.Label(
                        edit_title_frame,
                        text="修改供应商信息",
                        font=Styles.SUB_HEADER_FONT,
                        bg=Styles.BACKGROUND_COLOR,
                        fg=Styles.HEADER_COLOR
                    ).pack()

                    form_frame = tk.Frame(edit_top, bg=Styles.BACKGROUND_COLOR)
                    form_frame.pack(pady=Styles.PADY_MEDIUM)

                    vars_dict = {}
                    fields = [
                        ("供应商名称", "供应商名称", supplier_row['供应商名称']),
                        ("联系人", "联系人", supplier_row['联系人']),
                        ("联系电话", "联系电话", supplier_row['联系电话']),
                        ("地址", "地址", supplier_row['地址']),
                        ("备注", "备注", supplier_row['备注'])
                    ]

                    for i, (label, key, value) in enumerate(fields):
                        tk.Label(form_frame, text=label, font=Styles.LABEL_FONT, bg=Styles.BACKGROUND_COLOR, fg=Styles.TEXT_COLOR).grid(row=i, column=0, sticky='w', pady=5)
                        var = tk.StringVar(value=str(value))
                        vars_dict[key] = var
                        tk.Entry(form_frame, textvariable=var, width=30, font=Styles.TEXT_FONT).grid(row=i, column=1, pady=5)

                    def save():
                        updates = {}
                        for key, var in vars_dict.items():
                            value = var.get().strip()
                            updates[key] = value if value else ""

                        if updates:
                            all_suppliers = parent.system.excel_manager.get_all_suppliers()
                            idx = all_suppliers[all_suppliers['供应商编号'] == supplier_id].index
                            if len(idx) > 0:
                                for key, value in updates.items():
                                    if value:
                                        all_suppliers.at[idx[0], key] = value
                                parent.system.excel_manager.write_sheet("供应商", all_suppliers)
                                messagebox.showinfo("成功", "供应商信息更新成功！")
                                edit_top.destroy()
                                top.destroy()
                                parent.view_all_suppliers()
                            else:
                                messagebox.showerror("错误", "更新失败！")
                        else:
                            messagebox.showinfo("提示", "未做任何修改。")

                    btn_frame = tk.Frame(edit_top, bg=Styles.BACKGROUND_COLOR)
                    btn_frame.pack(pady=Styles.PADY_MEDIUM)

                    btn_save = tk.Button(btn_frame, text="保存修改", font=Styles.BUTTON_FONT,
                              width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT, command=save,
                              bg=Styles.PRIMARY_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
                    btn_save.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)
                    btn_save.bind("<Enter>", lambda e, b=btn_save: b.config(bg=Styles.BUTTON_HOVER_COLOR))
                    btn_save.bind("<Leave>", lambda e, b=btn_save: b.config(bg=Styles.PRIMARY_COLOR))

                    btn_cancel = tk.Button(btn_frame, text="取消", font=Styles.BUTTON_FONT,
                              width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT, command=edit_top.destroy,
                              bg=Styles.ERROR_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
                    btn_cancel.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)

        elif '客户编号' in columns:
            customer_id_idx = columns.index('客户编号')
            customer_id = values[customer_id_idx]

            df_customer = parent.system.excel_manager.get_all_customers()
            if not df_customer.empty:
                customer = df_customer[df_customer['客户编号'] == customer_id]
                if not customer.empty:
                    customer_row = customer.iloc[0]

                    edit_top = parent._create_toplevel_with_size("dataframe_edit_customer", "medium", parent=top)
                    edit_top.title("修改客户信息")
                    edit_top.configure(bg=Styles.BACKGROUND_COLOR)
                    edit_top.resizable(True, True)

                    edit_title_frame = tk.Frame(edit_top, bg=Styles.BACKGROUND_COLOR)
                    edit_title_frame.pack(pady=Styles.PADY_MEDIUM)

                    tk.Label(
                        edit_title_frame,
                        text="修改客户信息",
                        font=Styles.SUB_HEADER_FONT,
                        bg=Styles.BACKGROUND_COLOR,
                        fg=Styles.HEADER_COLOR
                    ).pack()

                    form_frame = tk.Frame(edit_top, bg=Styles.BACKGROUND_COLOR)
                    form_frame.pack(pady=Styles.PADY_MEDIUM)

                    vars_dict = {}
                    fields = [
                        ("客户名称", "客户名称", customer_row['客户名称']),
                        ("联系电话", "联系电话", customer_row['联系电话']),
                        ("电子邮箱", "电子邮箱", customer_row['电子邮箱']),
                        ("地址", "地址", customer_row['地址']),
                        ("备注", "备注", customer_row['备注'])
                    ]

                    for i, (label, key, value) in enumerate(fields):
                        tk.Label(form_frame, text=label, font=Styles.LABEL_FONT, bg=Styles.BACKGROUND_COLOR, fg=Styles.TEXT_COLOR).grid(row=i, column=0, sticky='w', pady=5)
                        var = tk.StringVar(value=str(value) if pd.notna(value) else "")
                        vars_dict[key] = var
                        tk.Entry(form_frame, textvariable=var, width=30, font=Styles.TEXT_FONT).grid(row=i, column=1, pady=5)

                    def save():
                        updates = {}
                        for key, var in vars_dict.items():
                            value = var.get().strip()
                            updates[key] = value if value else ""

                        if updates:
                            all_customers = parent.system.excel_manager.get_all_customers()
                            idx = all_customers[all_customers['客户编号'] == customer_id].index
                            if len(idx) > 0:
                                for key, value in updates.items():
                                    if value:
                                        all_customers.at[idx[0], key] = value
                                parent.system.excel_manager.write_sheet("客户信息", all_customers)
                                messagebox.showinfo("成功", "客户信息更新成功！")
                                edit_top.destroy()
                                top.destroy()
                                parent.view_all_customers()
                            else:
                                messagebox.showerror("错误", "更新失败！")
                        else:
                            messagebox.showinfo("提示", "未做任何修改。")

                    btn_frame = tk.Frame(edit_top, bg=Styles.BACKGROUND_COLOR)
                    btn_frame.pack(pady=Styles.PADY_MEDIUM)

                    btn_save = tk.Button(btn_frame, text="保存修改", font=Styles.BUTTON_FONT,
                              width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT, command=save,
                              bg=Styles.PRIMARY_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
                    btn_save.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)
                    btn_save.bind("<Enter>", lambda e, b=btn_save: b.config(bg=Styles.BUTTON_HOVER_COLOR))
                    btn_save.bind("<Leave>", lambda e, b=btn_save: b.config(bg=Styles.PRIMARY_COLOR))

                    btn_cancel = tk.Button(btn_frame, text="取消", font=Styles.BUTTON_FONT,
                              width=Styles.BUTTON_WIDTH, height=Styles.BUTTON_HEIGHT, command=edit_top.destroy,
                              bg=Styles.ERROR_COLOR, fg="white", relief=tk.FLAT, padx=10, pady=5)
                    btn_cancel.pack(side=tk.LEFT, padx=Styles.PADX_SMALL)


    def _on_tree_double_click(event):
        region = tree.identify_region(event.x, event.y)
        if region == "separator":
            col_id = tree.identify_column(event.x)
            col_idx = int(col_id.replace("#", "")) - 1
            if 0 <= col_idx < len(columns):
                col_name = columns[col_idx]
                if _search_filtered_df is not None:
                    _auto_fit_single_column(_search_filtered_df, col_name, col_idx)
                elif _original_df is not None:
                    _auto_fit_single_column(_original_df, col_name, col_idx)
                else:
                    page_df, _ = _load_page_data(current_page_var.get())
                    _auto_fit_single_column(page_df, col_name, col_idx)
            return
        show_detail(event)

    tree.bind('<Double-1>', _on_tree_double_click)

    status_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    status_frame.pack(pady=Styles.PADY_SMALL, fill=tk.X)

    tk.Label(
        status_frame,
        text=f"共 {total_rows} 条记录",
        font=Styles.TEXT_FONT,
        bg=Styles.BACKGROUND_COLOR,
        fg="#666666"
    ).pack(padx=Styles.PADX_MEDIUM, anchor=tk.W)

    btn_frame = tk.Frame(top, bg=Styles.BACKGROUND_COLOR)
    btn_frame.pack(pady=Styles.PADY_MEDIUM)

    if '供应商编号' in columns:
        def delete_selected_supplier():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("提示", "请先选择要删除的供应商")
                return

            item = tree.item(selected[0])
            values = item['values']
            supplier_id_idx = columns.index('供应商编号')
            supplier_id = values[supplier_id_idx]
            supplier_name_idx = columns.index('供应商名称')
            supplier_name = values[supplier_name_idx]

            if messagebox.askyesno("确认", f"确定要删除供应商 '{supplier_name}' 吗？"):
                all_suppliers = parent.system.excel_manager.get_all_suppliers()
                new_suppliers = all_suppliers[all_suppliers['供应商编号'] != supplier_id]
                parent.system.excel_manager.write_sheet("供应商", new_suppliers)
                messagebox.showinfo("成功", "供应商删除成功！")
                top.destroy()
                parent.view_all_suppliers()

        tk.Button(
            btn_frame,
            text="删除选中供应商",
            font=Styles.BUTTON_FONT,
            width=Styles.BUTTON_WIDTH,
            height=Styles.BUTTON_HEIGHT,
            command=delete_selected_supplier,
            bg=Styles.ERROR_COLOR,
            fg="white",
            relief=tk.FLAT,
            padx=10,
            pady=5
        ).pack(side=tk.LEFT, padx=10)

    if '客户编号' in columns:
        def delete_selected_customer():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("提示", "请先选择要删除的客户")
                return

            item = tree.item(selected[0])
            values = item['values']
            customer_id_idx = columns.index('客户编号')
            customer_id = values[customer_id_idx]
            customer_name_idx = columns.index('客户名称')
            customer_name = values[customer_name_idx]

            if messagebox.askyesno("确认", f"确定要删除客户 '{customer_name}' 吗？"):
                all_customers = parent.system.excel_manager.get_all_customers()
                new_customers = all_customers[all_customers['客户编号'] != customer_id]
                parent.system.excel_manager.write_sheet("客户信息", new_customers)
                messagebox.showinfo("成功", "客户删除成功！")
                top.destroy()
                parent.view_all_customers()

        tk.Button(
            btn_frame,
            text="删除选中客户",
            font=Styles.BUTTON_FONT,
            width=Styles.BUTTON_WIDTH,
            height=Styles.BUTTON_HEIGHT,
            command=delete_selected_customer,
            bg=Styles.ERROR_COLOR,
            fg="white",
            relief=tk.FLAT,
            padx=10,
            pady=5
        ).pack(side=tk.LEFT, padx=10)

    tk.Button(
        btn_frame,
        text="关闭",
        font=Styles.BUTTON_FONT,
        width=Styles.BUTTON_WIDTH,
        height=Styles.BUTTON_HEIGHT,
        command=top.destroy,
        bg=Styles.ERROR_COLOR,
        fg="white",
        relief=tk.FLAT,
        padx=10,
        pady=5
    ).pack(side=tk.LEFT, padx=10)